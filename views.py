from rest_framework import viewsets
from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes,authentication_classes
from rest_framework.response import Response
from .serializers import (
    RegistrarColaboradorSerializer,
    RegistrarEquipoSerializer,
    RegistrarLicenciaSerializer,
    MantenimientoSerializer,
    ImpresoraSerializer,
    RegistrarMantenImpreSerializer,
    RegistrarPerifericoSerializer,
    RegistrarUsuarioSerializer,
    userSerializer,
    ContrasenaSiesaSerializer, ContrasenaAntivirusSerializer, ContrasenaVPNSerializer,
    ContrasenaServidorSerializer, ContrasenaEquipoSerializer,EmpresaSerializer,CalendarioEventoSerializer,RegistroTareaSerializer
)
from .models import MantenimientoImpresora, Perifericos, RegistrarColaborador, RegistrarEquipo, RegistrarLicencia,Mantenimiento, Impresora, RegistrarUsuario,ContrasenaSiesa, ContrasenaAntivirus, ContrasenaVPN, ContrasenaServidor, ContrasenaEquipo,Empresa,CalendarioEvento,RegistroTarea
from rest_framework.parsers import MultiPartParser,FormParser
from rest_framework.authtoken.models import Token
from django.shortcuts import get_object_or_404
from rest_framework.permissions import AllowAny
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.contrib.auth.models import User
from django.contrib.auth import authenticate
from rest_framework.views import APIView
from aplicacion.models import RegistrarUsuario  # Asegúrate de importar tu modelo personalizado
from rest_framework import generics, permissions, status
from rest_framework.exceptions import AuthenticationFailed
from rest_framework.exceptions import PermissionDenied
from .views_base import EmpresaFilteredListCreateView, EmpresaFilteredRetrieveUpdateDestroyView, AdminOnlyMixin
from .middleware import get_empresa_actual
from captcha.models import CaptchaStore
from captcha.helpers import captcha_image_url


@api_view(['POST'])
@permission_classes([AllowAny])  # Permitir acceso sin autenticación
def register(request):
    serializer = userSerializer(data=request.data)
    if serializer.is_valid():
        # Crear usuario pero sin autenticar
        user = serializer.save()
        user.set_password(request.data['password'])  # Hashear contraseña
        user.save()
        return Response({
            'message': 'Usuario registrado correctamente. Por favor inicia sesión.',
            'user': serializer.data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    username = request.data.get('username')
    password = request.data.get('password')

    user = authenticate(username=username, password=password)
    if not user:
        return Response({'error': 'Credenciales inválidas'}, status=status.HTTP_401_UNAUTHORIZED)

    # Verificar que el usuario esté activo
    if not user.is_active:
        return Response({'error': 'Usuario inactivo'}, status=status.HTTP_401_UNAUTHORIZED)

    token, created = Token.objects.get_or_create(user=user)
    

    response_data = {
        'token': token.key,
        'user': {
            'username': user.username,
            'is_staff': user.is_staff
        }
    }
    return Response(response_data, status=status.HTTP_200_OK)



@api_view(['GET'])
@permission_classes([AllowAny])
def obtener_captcha(request):
    """
    Genera un captcha nuevo y devuelve el ID y la URL de la imagen
    """
    new_captcha = CaptchaStore.generate_key()
    url = captcha_image_url(new_captcha)
    return Response({'captcha_key': new_captcha, 'captcha_image_url': url})

@api_view(['POST'])
@permission_classes([AllowAny])
def cambiar_password_con_captcha(request):
    """
    Cambia la contraseña de un usuario verificando el captcha
    """
    username = request.data.get('username')
    new_password = request.data.get('password')
    captcha_key = request.data.get('captcha_key')
    captcha_value = request.data.get('captcha_value')

    if not all([username, new_password, captcha_key, captcha_value]):
        return Response({'error': 'Faltan campos'}, status=status.HTTP_400_BAD_REQUEST)

    # validar captcha
    from captcha.models import CaptchaStore
    try:
        captcha = CaptchaStore.objects.get(hashkey=captcha_key)
        if captcha.response.lower() != captcha_value.lower():
            return Response({'error': 'Captcha incorrecto'}, status=status.HTTP_400_BAD_REQUEST)
        captcha.delete()  # eliminar captcha usado
    except CaptchaStore.DoesNotExist:
        return Response({'error': 'Captcha inválido'}, status=status.HTTP_400_BAD_REQUEST)

    # cambiar contraseña
    try:
        user = RegistrarUsuario.objects.get(username=username)
        user.set_password(new_password)
        user.save()
        return Response({'status': 'Contraseña actualizada correctamente'}, status=status.HTTP_200_OK)
    except RegistrarUsuario.DoesNotExist:
        return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)



@api_view(['POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def logout(request):
    try:
        # Eliminar token de autenticación
        request.user.auth_token.delete()
    except (AttributeError, Token.DoesNotExist):
        pass
    
    # Limpiar respuesta
    response = Response({'message': 'Sesión cerrada correctamente'})
    response.delete_cookie('token')  # Si usas cookies
    return response

@api_view(['GET', 'POST'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def obtener_empresas_usuario(request):
    from .serializers import EmpresaSerializer
    
    if request.method == 'GET':
        # Listar todas las empresas activas
        empresas = Empresa.objects.filter(activa=True)
        serializer = EmpresaSerializer(empresas, many=True)
        return Response(serializer.data)

    if request.method == 'POST':
        empresa_id = request.data.get('empresa_id')
        try:
            empresa = Empresa.objects.get(id=empresa_id, activa=True)
        except Empresa.DoesNotExist:
            return Response({'error': 'Empresa no encontrada'}, status=404)

        request.session['empresa_actual_id'] = empresa.id  

        return Response({
            'message': 'Empresa seleccionada correctamente',
            'empresa_actual': EmpresaSerializer(empresa).data
        })


class RegistroTareaView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = RegistroTareaSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistroTarea.objects.all()

    def update(self, request, *args, **kwargs):
        print("Datos recibidos para update:", request.data)  # debug opcional
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


class RegistroTareaDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistroTareaSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistroTarea.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Registro eliminado correctamente"}, status=status.HTTP_204_NO_CONTENT)
    
class RegistrarEquipoView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = RegistrarEquipoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarEquipo.objects.all()
    
    def update(self, request, *args, **kwargs):
        print("Datos recibidos:", request.data)  # Para debug
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)

class RegistrarEquipoDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistrarEquipoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarEquipo.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Equipo eliminado correctamente"}, status=status.HTTP_204_NO_CONTENT)

@api_view(['GET'])
def equipo_por_colaborador(request, id_colaborador):
    try:
        empresa_actual=get_empresa_actual(request)

        if not empresa_actual:
            return Response({'error': 'No se pudo determinar la empresa actual'}, status=404)
        
        equipos = RegistrarEquipo.objects.filter(responsable_id=id_colaborador, empresa=empresa_actual)

        if equipos.exists():
            serializer = RegistrarEquipoSerializer(equipos,many=True, context={'request': request})
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(
                {"error": "No se encontró ningún equipo para este colaborador"},
                status=status.HTTP_404_NOT_FOUND
            )
    except Exception as e:
        return Response(
            {"error": f"Error del servidor: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def impresora_con_mantenimiento(request, id_impresora):
    try:
        empresa_actual = get_empresa_actual(request) 

        if not empresa_actual:
            return Response(
                {"error": "No se pudo determinar la empresa actual"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Buscar impresora por ID y empresa actual
        impresora = Impresora.objects.filter(
            id=id_impresora,
            empresa=empresa_actual
        ).first()

        if not impresora:
            return Response(
                {"error": "No se encontró ninguna impresora con este ID en la empresa actual"},
                status=status.HTTP_404_NOT_FOUND
            )

        # Obtener mantenimientos relacionados en la misma empresa
        mantenimientos = MantenimientoImpresora.objects.filter(
            impresora=impresora,
            empresa=empresa_actual
        )

        return Response({
            "impresora": ImpresoraSerializer(
                impresora, context={'request': request}
            ).data,
            "mantenimientos": RegistrarMantenImpreSerializer(
                mantenimientos, many=True
            ).data
        })
    except Exception as e:
        return Response(
            {"error": f"Error del servidor: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

            
    
        
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def generar_reporte_usuarios_por_area(request):
    # Obtener el área desde los parámetros de la solicitud
    area = request.GET.get('area')

    if not area:
        return HttpResponse("Debes proporcionar un área.", status=400)

    # Filtrar los usuarios por área
    empresa_actual=get_empresa_actual(request)
    usuarios = RegistrarColaborador.objects.filter(
        area=area,
        empresa=empresa_actual 
    )

    if not usuarios.exists():
        return HttpResponse("No se encontraron usuarios para el área especificada.", status=404)

    # Crear un libro de Excel
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Usuarios y Equipos por Área"

    # Agregar encabezados
    worksheet.append([
        "ID Usuario", "Nombre", "Apellido", "Área", "Cargo", "Empresa",
        # Datos del usuario
        "ID Equipo", "Marca", "Modelo", "Memoria", "Procesador", "Office", "Serial", "Sistema Operativo", "Fecha Adquisición", "Estado"  
    ])

    # Agregar datos de los usuarios y sus equipos
    for usuario in usuarios:
        # Obtener el equipo asociado al usuario (suponiendo una relación uno a uno)
        equipo = RegistrarEquipo.objects.filter(responsable=usuario).first()

        worksheet.append([
            # Datos del usuario
            usuario.id,
            usuario.nombre,
            usuario.apellido,
            usuario.area,
            usuario.cargo,
            str(usuario.empresa),
            # Datos del equipo (si existe)
            equipo.id if equipo else "N/A",
            equipo.marca if equipo else "N/A",
            equipo.modelo if equipo else "N/A",
            equipo.memoria if equipo else "N/A",
            equipo.procesador if equipo else "N/A",
            equipo.office if equipo else "N/A",
            equipo.serial if equipo else "N/A",
            equipo.sistema_operativo if equipo else "N/A",
            equipo.fecha_adquisicion if equipo else "N/A",
            equipo.estado if equipo else "N/A",
        ])

    for col in worksheet.columns:
        max_length = 0
        col_letter=get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width
          

    # Crear una respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=usuarios_y_equipos_{area}.xlsx'

    # Guardar el libro de Excel en la respuesta
    workbook.save(response)

    return response

@api_view(['GET'])
def obtener_licencias_por_equipo(request, equipo_id):
    licencias = RegistrarLicencia.objects.filter(equipo_id=equipo_id).values(
        'tipo_licencia'
    )
    return Response(list(licencias))

class RegistrarUsuarioViewSet(AdminOnlyMixin, generics.ListCreateAPIView):
     queryset = RegistrarUsuario.objects.all()
     serializer_class = RegistrarUsuarioSerializer
     permission_classes = [permissions.IsAuthenticated]

class RegistrarUsuarioDetailView(AdminOnlyMixin, generics.RetrieveUpdateDestroyAPIView):
     queryset = RegistrarUsuario.objects.all()
     serializer_class = RegistrarUsuarioSerializer
     permission_classes = [permissions.IsAuthenticated]


class RegistrarColaboradorView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = RegistrarColaboradorSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarColaborador.objects.all()

    def perform_create(self, serializer):
        serializer.save()
 

class RegistrarColaboradorDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistrarColaboradorSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarColaborador.objects.all()

class RegistrarColaboradorAllView(generics.ListAPIView):
    """
    Lista de todos los colaboradores, sin filtro por empresa
    (para asignar responsables libremente en equipos).
    """
    serializer_class = RegistrarColaboradorSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarColaborador.objects.all()    



class RegistrarLicenciaView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = RegistrarLicenciaSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarLicencia.objects.all()
    
    def update(self, request, *args, **kwargs):
        print("Datos recibidos:", request.data)  # Para debug
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)         


class RegistrarLicenciaDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistrarLicenciaSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = RegistrarLicencia.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Licencia eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)

class MantenimientoView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = MantenimientoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Mantenimiento.objects.all()
    
    def update(self, request, *args, **kwargs):
        print("Datos recibidos:", request.data)  # Para debug
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)    
    
class MantenimientoDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = MantenimientoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Mantenimiento.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Mantenimiento eliminado correctamente"}, status=status.HTTP_204_NO_CONTENT)
    
 # Vista para licencias por equipo
class LicenciasPorEquipoView(generics.ListAPIView):
    serializer_class = RegistrarLicenciaSerializer
    
    def get_queryset(self):
        equipo_id = self.kwargs['equipo_id']
        return RegistrarLicencia.objects.filter(equipo_id=equipo_id)

# Vista para mantenimientos por equipo  
class MantenimientosPorEquipoView(generics.ListAPIView):
    serializer_class = MantenimientoSerializer
    
    def get_queryset(self):
        equipo_id = self.kwargs['equipo_id']
        return Mantenimiento.objects.filter(equipo_id=equipo_id)   
    

class CalendarioEventoView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = CalendarioEventoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CalendarioEvento.objects.all()

    
    
    def update(self, request, *args, **kwargs):
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)


class CalendarioEventoDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = CalendarioEventoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = CalendarioEvento.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Evento eliminado correctamente"}, status=status.HTTP_204_NO_CONTENT)


class RegistrarImpresoraView(EmpresaFilteredListCreateView):
    serializer_class = ImpresoraSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Impresora.objects.all()


class RegistrarImpresoraDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ImpresoraSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Impresora.objects.all()
    

class RegistrarPerifericoView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = RegistrarPerifericoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Perifericos.objects.all()

   
    
    def update(self, request, *args, **kwargs):
        print("Datos recibidos:", request.data)  # Para debug
        return super().update(request, *args, **kwargs)

    def patch(self, request, *args, **kwargs):
        return self.partial_update(request, *args, **kwargs)  


class RegistrarPerifericoDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistrarPerifericoSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Perifericos.objects.all()

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Periférico eliminado correctamente"}, status=status.HTTP_204_NO_CONTENT)


class MantenimpreView(EmpresaFilteredListCreateView):
    serializer_class = RegistrarMantenImpreSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = MantenimientoImpresora.objects.all()
    


class MantenimpreDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = RegistrarMantenImpreSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = MantenimientoImpresora.objects.all() 


class ContrasenaSiesaView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = ContrasenaSiesaSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaSiesa.objects.all()

    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({"message": "Contraseña Siesa creada correctamente", "data": response.data}, status=status.HTTP_201_CREATED)

class ContrasenaSiesaDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ContrasenaSiesaSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaSiesa.objects.all()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({"message": "Contraseña Siesa actualizada correctamente", "data": response.data}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Contraseña Siesa eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)

# Antivirus
class ContrasenaAntivirusView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = ContrasenaAntivirusSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaAntivirus.objects.all() 

    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({"message": "Contraseña Antivirus creada correctamente", "data": response.data}, status=status.HTTP_201_CREATED)

class ContrasenaAntivirusDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ContrasenaAntivirusSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaAntivirus.objects.all()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({"message": "Contraseña Antivirus actualizada correctamente", "data": response.data}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Contraseña Antivirus eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)

# VPN
class ContrasenaVPNView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = ContrasenaVPNSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaVPN.objects.all()
    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({"message": "Contraseña VPN creada correctamente", "data": response.data}, status=status.HTTP_201_CREATED)

class ContrasenaVPNDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ContrasenaVPNSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaVPN.objects.all()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({"message": "Contraseña VPN actualizada correctamente", "data": response.data}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs): 
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Contraseña VPN eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)

# Servidor
class ContrasenaServidorView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = ContrasenaServidorSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaServidor.objects.all()
    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({"message": "Contraseña Servidor creada correctamente", "data": response.data}, status=status.HTTP_201_CREATED)

class ContrasenaServidorDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ContrasenaServidorSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaServidor.objects.all()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({"message": "Contraseña Servidor actualizada correctamente", "data": response.data}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Contraseña Servidor eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)

# Equipo
class ContrasenaEquipoView(AdminOnlyMixin, EmpresaFilteredListCreateView):
    serializer_class = ContrasenaEquipoSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaEquipo.objects.all()
    
    def create(self, request, *args, **kwargs):
        response = super().create(request, *args, **kwargs)
        return Response({"message": "Contraseña Equipo creada correctamente", "data": response.data}, status=status.HTTP_201_CREATED)

class ContrasenaEquipoDetailView(AdminOnlyMixin, EmpresaFilteredRetrieveUpdateDestroyView):
    serializer_class = ContrasenaEquipoSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes=(MultiPartParser,FormParser)
    queryset = ContrasenaEquipo.objects.all()

    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        return Response({"message": "Contraseña Equipo actualizada correctamente", "data": response.data}, status=status.HTTP_200_OK)

    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        self.perform_destroy(instance)
        return Response({"message": "Contraseña Equipo eliminada correctamente"}, status=status.HTTP_204_NO_CONTENT)


    
# Funciones para obtener totales filtrados por empresa
from django.http import JsonResponse

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_equipos(request):
    empresa_actual = get_empresa_actual(request)
    total = RegistrarEquipo.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total": total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_colaboradores(request):
    empresa_actual = get_empresa_actual(request)
    total = RegistrarColaborador.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_licencias(request):
    empresa_actual = get_empresa_actual(request)
    total = RegistrarLicencia.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_mantenimientos(request):
    empresa_actual = get_empresa_actual(request)
    total = Mantenimiento.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_impresoras(request):
    empresa_actual = get_empresa_actual(request)
    total = Impresora.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_perifericos(request):
    empresa_actual = get_empresa_actual(request)
    total = Perifericos.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_contrasena_siesa(request):
    empresa_actual = get_empresa_actual(request)
    total = ContrasenaSiesa.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_contrasena_antivirus(request):
    empresa_actual = get_empresa_actual(request)
    total = ContrasenaAntivirus.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_contrasena_vpn(request):
    empresa_actual = get_empresa_actual(request)
    total = ContrasenaVPN.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_contrasena_servidor(request):
    empresa_actual = get_empresa_actual(request)
    total = ContrasenaServidor.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def total_contrasena_equipo(request):
    empresa_actual = get_empresa_actual(request)
    total = ContrasenaEquipo.objects.filter(empresa=empresa_actual).count() if empresa_actual else 0
    return JsonResponse({"total":total})


@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def obtener_empresa_actual(request):
    empresa_actual= get_empresa_actual(request)

    if not empresa_actual:
        return Response({'error': 'No hay empresa seleccionada'}, status=404)
    
    from .serializers import EmpresaSerializer

    serializer = EmpresaSerializer(empresa_actual)
    return Response(serializer.data)

@api_view(['GET'])
@authentication_classes([TokenAuthentication])
@permission_classes([IsAuthenticated])
def obtener_datos_empresa_filtrados(request):
   
    empresa_actual =get_empresa_actual(request)
    
    if not empresa_actual:
        return Response({'error': 'No hay empresa seleccionada'}, status=404)
    
    # Obtener datos filtrados por empresa
    datos = {
        'equipos': RegistrarEquipo.objects.filter(empresa=empresa_actual).count(),
        'colaboradores': RegistrarColaborador.objects.filter(empresa=empresa_actual).count(),
        'licencias': RegistrarLicencia.objects.filter(empresa=empresa_actual).count(),
        'mantenimientos': Mantenimiento.objects.filter(empresa=empresa_actual).count(),
        'impresoras': Impresora.objects.filter(empresa=empresa_actual).count(),
        'perifericos': Perifericos.objects.filter(empresa=empresa_actual).count(),
        'contrasenas_siesa': ContrasenaSiesa.objects.filter(empresa=empresa_actual).count(),
        'contrasenas_antivirus': ContrasenaAntivirus.objects.filter(empresa=empresa_actual).count(),
        'contrasenas_vpn': ContrasenaVPN.objects.filter(empresa=empresa_actual).count(),
        'contrasenas_servidor': ContrasenaServidor.objects.filter(empresa=empresa_actual).count(),
        'contrasenas_equipo': ContrasenaEquipo.objects.filter(empresa=empresa_actual).count(),
    }
    
    return Response(datos)